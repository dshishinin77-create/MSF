from openpyxl import load_workbook
import re
import os
import sys
import datetime
import json
import traceback


def smart_load_workbook(file_path):
    # data_only=True ускоряет чтение и берет только значения (без формул)
    return load_workbook(file_path, data_only=True)


def re_var(consts: list, string: str, stop_words=None):
    string = str(string).replace('\n', ' ')
    reg = '.*' + '.*'.join(consts) + '.*'
    if stop_words:
        for word in stop_words:
            stop = '.*' + word + '.*'
            if bool(re.match(f'{stop}', string, re.IGNORECASE)):
                return False
    return bool(re.match(f'{reg}', string, re.IGNORECASE))


def header_in_reg(header, position, dict_of_reg):
    val = dict_of_reg[position]
    if type(val[0]) == str:
        return re_var(val, header)
    else:
        consts = val[0]
        stop_words = val[1] if len(val) > 1 else None
        return re_var(consts, header, stop_words)


def main_func(table_name):
    short_table_name = os.path.basename(table_name)
    print(f'\n>>> Обработка файла: {short_table_name}')

    try:
        wb = smart_load_workbook(table_name)
    except Exception as e:
        print(f"    [ОШИБКА] Не удалось открыть файл: {e}")
        return None

    # 1. Ищем нужную вкладку
    ws = None
    for sheet in wb.worksheets:
        if 'change request' in sheet.title.lower():
            ws = sheet
            print(f"    [OK] Найдена вкладка: '{sheet.title}'")
            break

    if ws is None:
        print(
            "    [ОШИБКА] Вкладка 'Change Request' не найдена! Пропускаем файл.")
        return None

    # 2. Словарь регулярных выражений для поиска ключей
    dict_of_keys = {
        'CR_number': [['change', 'request', 'no'], ['internal']],
        'Reg_date': ['registration', 'date'],
        'CR_coordinator': ['change', 'coordinator'],
        'Change_type': ['type', 'change'],
        'Constr_facility': ['construction', 'facility'],
        'Document_type': ['type', 'documentation'],
        'Organization': ['initiator', 'organization'],
        'Initiator': [['change', 'initiator'],
                      ['organization', 'internal', 'coordinator']],
        'Initiator_internal_CR': ['initiator', 'internal', 'cr'],  # Новое поле
        'CR_reason': [['change', 'reason'], ['&', 'description']],
        # Добавлены стоп-слова для пропуска серых заголовков
        'Descr_tech_sol': ['change', 'description'],
        'Evaluation': ['change', 'evaluation'],
        'SSC': ['building', 'kks'],
        'TDD_sets': ['tdd', 'code']
    }

    # Итоговый словарь, изначально заполненный None и пустыми словарями
    CR_d = {key: None for key in dict_of_keys.keys()}
    CR_d['SSC'] = {}
    CR_d['TDD_sets'] = {}

    max_col = ws.max_column
    max_row = ws.max_row

    print("    [Сканирование] Поиск значений...")

    # 3. Сканируем ячейки
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is None or str(cell_val).strip() == '':
                continue

            cell_text = str(cell_val).lower()

            # Проверяем, совпадает ли текст ячейки с каким-либо ключом
            matched_key = None
            for key in list(dict_of_keys.keys()):
                if header_in_reg(cell_text, key, dict_of_keys):
                    matched_key = key
                    break

            if matched_key:
                extracted_value = None

                # Ищем ближайшую непустую ячейку справа в этой же строке
                for c in range(col + 1, max_col + 1):
                    neighbor_val = ws.cell(row=row, column=c).value
                    if neighbor_val is not None and str(
                            neighbor_val).strip() != '':

                        # Защита: проверяем, не является ли найденный текст следующим заголовком
                        is_header = False
                        for k in dict_of_keys.keys():
                            if header_in_reg(str(neighbor_val).lower(), k,
                                             dict_of_keys):
                                is_header = True
                                break

                        if not is_header:
                            extracted_value = neighbor_val
                        break  # Нашли текст (значение или другой заголовок) — дальше вправо не идем

                print(
                    f"      -> Найдено: {matched_key} | Значение: {extracted_value}")

                # Создаем отдельные блоки для SSC и TDD_sets
                if matched_key in ['SSC', 'TDD_sets']:
                    if extracted_value:
                        val_str = str(extracted_value).strip()
                        # Разбиваем по переносам строки внутри ячейки Excel
                        lines = [line.strip() for line in val_str.split('\n')
                                 if line.strip()]
                        CR_d[matched_key] = {line: {} for line in lines}
                else:
                    CR_d[matched_key] = extracted_value

                dict_of_keys.pop(matched_key, None)  # Удаляем найденный ключ

    print(f'>>> Файл {short_table_name} успешно разобран.')
    return CR_d


def js_ser(obj):
    if isinstance(obj, datetime.datetime):
        return obj.strftime('%d.%m.%Y')
    return str(obj)


def output(option):
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    if option == 0:
        excel_path = 'D:\\!Digital_twin\\!CR\\CR_parser\\cr_test\\18_05\\'
    else:
        if option == 'pwd':
            excel_path = base_dir + '\\'
        else:
            user_input = input(
                'Введите ПОЛНЫЙ путь к папке с файлами (или нажмите Enter для текущей папки): ')
            if user_input.strip() == '':
                excel_path = base_dir + '\\'
            else:
                excel_path = user_input
                if excel_path and excel_path[-1] != '\\':
                    excel_path += '\\'

    print(f"\nИщем файлы Excel в папке: {excel_path}")

    try:
        all_files = os.listdir(path=excel_path)
    except Exception as e:
        print(f"Ошибка при доступе к папке: {e}")
        return

    list_of_tables = [x for x in all_files if (
                x.lower().endswith('.xls') or x.lower().endswith(
            '.xlsx')) and not x.startswith('~$')]

    print(f"Найдено файлов к обработке: {len(list_of_tables)}")
    if len(list_of_tables) == 0:
        return

    for file in list_of_tables:
        full_path = os.path.join(excel_path, file)
        result_dict = main_func(full_path)

        if result_dict:
            short_name = os.path.splitext(file)[0]
            out_file = os.path.join(excel_path, f"{short_name}.txt")

            with open(out_file, 'w', encoding='utf-8') as f:
                json.dump(result_dict, f, default=js_ser, ensure_ascii=False,
                          indent=2)

            print(f'>>> Готово! Результат сохранен в: {short_name}.txt')


if __name__ == '__main__':
    try:
        output('pwd')
    except Exception as e:
        print(f"\nПроизошла критическая ошибка:")
        traceback.print_exc()
    finally:
        input("\nРабота завершена. Нажмите Enter, чтобы выйти...")